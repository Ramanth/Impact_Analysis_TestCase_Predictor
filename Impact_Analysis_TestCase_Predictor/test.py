from sklearn.multiclass import OneVsRestClassifier;
from sklearn.multiclass import OneVsOneClassifier;
from sklearn.neighbors import KNeighborsClassifier
from sklearn.svm import LinearSVC;
from sklearn.neural_network import MLPClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier, AdaBoostClassifier;

from sklearn.preprocessing import MultiLabelBinarizer;
from openpyxl import load_workbook

workbook_path = "impact.xlsx";
wb = load_workbook(workbook_path);

# grab the active worksheet
ws = wb['Sheet1']
_X = [];
y = [];

def getBinaryVector(input):
    vec = [];
    arr = ['Login','Authentication','HomePage','Page1','Page2','Page3','CommonFn_12','CommonFn_13','CommonFn_23','Logout'];
    for i in arr:
        if i in input:
            vec.append(1);
        else:
            vec.append(0);

    return vec;

def createFeatureVector(input):
    vector = [];
    for eachInput in input:
        v = getBinaryVector(eachInput)
        vector.append(v);
    return vector;

for _column in range (3,17):
    _X.append(ws.cell(column =  _column, row = 27).value.split('\n'));

for _column in range (3,17):
    y.append(ws.cell(column =  _column, row = 28).value.split('\n'));

# Data can be assigned directly to cells

X = createFeatureVector(_X);

_X_Predict = [];
_X_Predict.append(ws.cell(column =  17, row = 27).value.split('\n'));

X_Predict = createFeatureVector(_X_Predict);

print(_X_Predict)

y_train = MultiLabelBinarizer().fit_transform(y);

m = OneVsRestClassifier(DecisionTreeClassifier());
m.fit(X, y_train);

output = m.predict(X_Predict);

bitarray = m.label_binarizer_.inverse_transform(output, threshold = 0.5);
classnames = MultiLabelBinarizer().fit(y).inverse_transform(bitarray);

print(classnames);

# Save the file
wb.save(workbook_path);